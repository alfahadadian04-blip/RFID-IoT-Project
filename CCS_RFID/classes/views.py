import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_GET, require_POST
from django.utils import timezone
from datetime import datetime, time, timedelta
from .models import Class, Enrollment, ClassSession, Attendance, ClassPDFReport
from CCS.models import User, PendingRFID
from CCS.forms import AdminLoginForm, StudentLoginForm, AdminRegistrationForm, StudentRegistrationForm
import traceback
import json
import logging
import re
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import pytz

logger = logging.getLogger(__name__)


# ============================================
# Validation function for Student ID and Email
# ============================================
def validate_student_registration_data(data):
    """Validate Student ID format and Email domain before form submission"""
    errors = {}
    
    student_id = data.get('student_id')
    if student_id:
        student_id_pattern = r'^\d{4}-\d{5}$'
        if not re.match(student_id_pattern, student_id):
            errors['student_id'] = 'Student ID must be in format: YYYY-XXXXX (e.g., 2022-00779)'
    
    email = data.get('email')
    if email:
        if not email.lower().endswith('@wmsu.edu.ph'):
            errors['email'] = 'Only @wmsu.edu.ph email addresses are allowed'
    
    return errors


def get_cell_value(cell):
    """Safely get value from a cell"""
    if cell is None:
        return None
    if hasattr(cell, 'value'):
        return cell.value
    return cell


def parse_excel_time(value):
    """Convert Excel time (datetime, float, or string) to time object"""
    if value is None:
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, (int, float)):
        total_seconds = int(value * 86400)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return time(hours, minutes, seconds)
    if isinstance(value, str):
        value = value.strip()
        for fmt in ['%I:%M%p', '%I:%M %p', '%H:%M', '%I:%M%p']:
            try:
                return datetime.strptime(value, fmt).time()
            except:
                continue
    return None


# ============================================
# RFID HANDLER (for Arduino)
# ============================================
@csrf_exempt
def rfid_handler(request):
    """API endpoint that receives RFID data from Arduino"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rfid_tag = data.get('rfid_tag')
            student_id = data.get('student_id')
            
            if not rfid_tag:
                return JsonResponse({'error': 'No RFID tag provided'}, status=400)
            
            print(f"📇 RFID received: {rfid_tag}, Student ID: {student_id}")
            
            if student_id:
                try:
                    student = User.objects.get(id=student_id, user_type='student')
                    existing = User.objects.filter(rfid_tag=rfid_tag).exclude(id=student_id).first()
                    if existing:
                        return JsonResponse({
                            'status': 'error',
                            'message': 'RFID tag already registered to another student'
                        }, status=400)
                    
                    student.rfid_tag = rfid_tag
                    student.save()
                    
                    # CLEAR THE PENDING REGISTRATION AFTER SUCCESSFUL REGISTRATION
                    PendingRFID.objects.filter(student=student).delete()
                    
                    print(f"✅ RFID registered for student: {student.get_full_name()}")
                    return JsonResponse({
                        'status': 'success',
                        'action': 'registered',
                        'name': student.get_full_name(),
                        'rfid_tag': rfid_tag
                    })
                    
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Student not found'}, status=404)
            
            else:
                student = User.objects.filter(
                    rfid_tag=rfid_tag, 
                    user_type='student',
                    is_active=True
                ).first()
                
                if student:
                    print(f"✅ Attendance recorded for: {student.get_full_name()}")
                    return JsonResponse({
                        'status': 'success',
                        'name': student.get_full_name(),
                        'action': 'attendance_recorded'
                    })
                else:
                    print(f"❌ No student found with RFID: {rfid_tag}")
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Student not found'
                    }, status=404)
                
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            print(f"❌ RFID handler error: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ============================================
# RFID PENDING REGISTRATION
# ============================================
@login_required
def register_rfid_for_student(request, student_id):
    """Assign RFID tag to a student during registration"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rfid_tag = data.get('rfid_tag')
            
            if not rfid_tag:
                return JsonResponse({'error': 'No RFID tag provided'}, status=400)
            
            if User.objects.filter(rfid_tag=rfid_tag).exists():
                return JsonResponse({
                    'error': 'RFID tag already registered to another student'
                }, status=400)
            
            student = User.objects.get(id=student_id, user_type='student')
            student.rfid_tag = rfid_tag
            student.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'RFID registered successfully'
            })
            
        except User.DoesNotExist:
            return JsonResponse({'error': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def check_rfid_status(request, student_id):
    """Check if an RFID has been registered for this student"""
    try:
        student = User.objects.get(id=student_id, user_type='student')
        
        if student.rfid_tag:
            return JsonResponse({
                'rfid_detected': True,
                'rfid_tag': student.rfid_tag
            })
        else:
            return JsonResponse({
                'rfid_detected': False
            })
    except User.DoesNotExist:
        return JsonResponse({'error': 'Student not found'}, status=404)


def create_pending_rfid(request, student_id):
    """Create a pending RFID registration for a student"""
    if request.method == 'POST':
        try:
            student = User.objects.get(id=student_id, user_type='student')
            PendingRFID.objects.filter(student=student).delete()
            pending = PendingRFID.objects.create(
                student=student,
                expires_at=timezone.now() + timedelta(minutes=5)
            )
            
            return JsonResponse({
                'success': True,
                'expires_at': pending.expires_at.isoformat()
            })
        except User.DoesNotExist:
            return JsonResponse({'error': 'Student not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def check_pending_rfid(request):
    """Check if any student is waiting for RFID registration"""
    PendingRFID.objects.filter(expires_at__lt=timezone.now()).delete()
    pending = PendingRFID.objects.first()
    
    if pending and not pending.is_expired():
        return JsonResponse({
            'waiting': True,
            'student_id': pending.student.id,
            'student_name': pending.student.get_full_name(),
            'expires_at': pending.expires_at.isoformat()
        })
    else:
        return JsonResponse({'waiting': False})

@csrf_exempt
def clear_pending_rfid(request):
    """Clear pending RFID registration (called when registration is cancelled or completed)"""
    if request.method == 'POST':
        try:
            # Delete all expired and pending registrations
            PendingRFID.objects.filter(expires_at__lt=timezone.now()).delete()
            
            # If specific student_id is provided, delete that pending registration
            data = json.loads(request.body) if request.body else {}
            student_id = data.get('student_id')
            
            if student_id:
                PendingRFID.objects.filter(student_id=student_id).delete()
            else:
                # Delete the oldest pending registration
                pending = PendingRFID.objects.first()
                if pending:
                    pending.delete()
            
            return JsonResponse({'success': True, 'message': 'Pending registration cleared'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

@csrf_exempt
def cancel_pending_registration(request):
    """Cancel pending registration for a specific student"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            student_id = data.get('student_id')
            
            if student_id:
                deleted_count, _ = PendingRFID.objects.filter(student_id=student_id).delete()
                return JsonResponse({
                    'success': True, 
                    'message': f'Registration cancelled for student ID: {student_id}',
                    'deleted': deleted_count
                })
            else:
                return JsonResponse({'error': 'Student ID required'}, status=400)
                
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ============================================
# AUTHENTICATION VIEWS
# ============================================
def adminLogin(request):
    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            if user.user_type == 'admin':
                login(request, user)
                messages.success(request, f'Welcome back, {user.first_name}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'This account is not an admin account.')
        else:
            messages.error(request, 'Invalid email or password.')
    
    return render(request, 'adminLogin.html', {})


def studentLogin(request):
    if request.method == 'POST':
        email = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=email, password=password)
        
        if user is not None:
            if user.user_type == 'student':
                login(request, user)
                messages.success(request, f'Welcome back, {user.first_name}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'This account is not a student account.')
        else:
            messages.error(request, 'Invalid email or password.')
    
    return render(request, 'studentLogin.html', {})


def adminRegistration(request):
    if request.method == 'POST':
        form = AdminRegistrationForm(request.POST)
        if form.is_valid():
            try:
                user = form.save()
                messages.success(request, 'Account created successfully! You can now login.')
                return redirect('adminLogin')
            except Exception as e:
                messages.error(request, f'Error creating account: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = AdminRegistrationForm()
    
    return render(request, 'adminRegistration.html', {'form': form})


def studentRegistration(request):
    if request.method == 'POST':
        print("=" * 50)
        print("Received POST data:")
        for key, value in request.POST.items():
            print(f"{key}: {value}")
        print("=" * 50)
        
        validation_errors = validate_student_registration_data(request.POST)
        
        if validation_errors:
            print("❌ Validation errors:", validation_errors)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                error_messages = []
                for field, error in validation_errors.items():
                    error_messages.append(f"{field}: {error}")
                return JsonResponse({
                    'success': False, 
                    'error': ' | '.join(error_messages)
                }, status=400)
            else:
                for field, error in validation_errors.items():
                    messages.error(request, error)
                return render(request, 'studentRegistration.html')
        
        student_id = request.POST.get('student_id')
        if student_id and User.objects.filter(student_id=student_id).exists():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Student ID already exists.'}, status=400)
            else:
                messages.error(request, 'Student ID already exists.')
                return render(request, 'studentRegistration.html')
        
        email = request.POST.get('email')
        if email and User.objects.filter(email=email).exists():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Email already exists.'}, status=400)
            else:
                messages.error(request, 'Email already exists.')
                return render(request, 'studentRegistration.html')
        
        form = StudentRegistrationForm(request.POST)
        
        if form.is_valid():
            try:
                user = form.save()
                print(f"✅ User created with ID: {user.id}")
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'student_id': user.id,
                        'message': 'Student created successfully'
                    })
                else:
                    messages.success(request, 'Student account created successfully!')
                    return render(request, 'studentRegistration.html', {'student_id': user.id})
                    
            except Exception as e:
                print(f"❌ Error saving user: {str(e)}")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)}, status=400)
                else:
                    messages.error(request, f'Error creating account: {str(e)}')
                    return render(request, 'studentRegistration.html')
        else:
            print("❌ Form is invalid!")
            print("Form errors:", form.errors)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                error_messages = []
                for field, errors in form.errors.items():
                    error_messages.append(f"{field}: {', '.join(errors)}")
                return JsonResponse({'success': False, 'error': ' | '.join(error_messages)}, status=400)
            else:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
                return render(request, 'studentRegistration.html')
    
    return render(request, 'studentRegistration.html', {})


# ============================================
# TEACHER CLASS VIEWS
# ============================================
@login_required
def dashboard(request):
    return render(request, 'dashboard.html', {'user': request.user})


@login_required
def attendance(request):
    return render(request, 'attendance.html')


@login_required
def classes(request):
    """Display all classes for the logged-in teacher"""
    user_classes = Class.objects.filter(teacher=request.user)
    total_students = sum(c.enrollments.count() for c in user_classes)
    
    context = {
        'classes': user_classes,
        'total_classes': user_classes.count(),
        'total_students': total_students,
        'avg_attendance': 87,
    }
    return render(request, 'classes.html', context)


@login_required
def view_class(request):
    """Display details of a specific class with student list"""
    class_id = request.GET.get('class_id')
    if not class_id:
        messages.error(request, 'No class specified.')
        return redirect('classes')
    
    try:
        class_obj = Class.objects.get(id=class_id, teacher=request.user)
    except Class.DoesNotExist:
        messages.error(request, 'Class not found.')
        return redirect('classes')
    
    enrollments = class_obj.enrollments.select_related('student').all()
    total_students = enrollments.count()
    
    students = []
    for idx, enrollment in enumerate(enrollments, 1):
        student = enrollment.student
        students.append({
            'number': idx,
            'student_id': student.student_id if hasattr(student, 'student_id') else '',
            'name': student.get_full_name(),
            'email': student.email,
            'gender': getattr(student, 'gender', 'N/A'),
            'status': enrollment.status,
            'absences': enrollment.absence_count,
            'course': getattr(student, 'course', ''),
        })
    
    context = {
        'class': class_obj,
        'students': students,
        'total_students': total_students,
        'present_today': 0,
        'absent_today': 0,
        'late_today': 0,
    }
    return render(request, 'view_class.html', context)


@login_required
def upload_masterlist(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        print(f"[DEBUG] File received: {excel_file.name}")
        
        try:
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = workbook.active
            
            print("=== FIRST 20 ROWS OF EXCEL ===")
            for i in range(1, 21):
                row = []
                for j in range(1, 9):
                    val = get_cell_value(sheet.cell(row=i, column=j))
                    row.append(val)
                print(f"Row {i}: {row}")
            
            school_year = None
            semester = None
            student_type = None
            for i in range(1, 15):
                val = get_cell_value(sheet.cell(row=i, column=1))
                if val and 'SCHOOL YEAR' in str(val).upper():
                    school_year = get_cell_value(sheet.cell(row=i, column=2))
                elif val and 'SEMESTER' in str(val).upper():
                    semester = get_cell_value(sheet.cell(row=i, column=2))
                elif val and 'STUDENT TYPE' in str(val).upper():
                    student_type = get_cell_value(sheet.cell(row=i, column=2))
            
            print(f"[DEBUG] Metadata - School Year: {school_year}, Semester: {semester}, Student Type: {student_type}")
            
            subject_start_row = None
            for i in range(1, 100):
                val = get_cell_value(sheet.cell(row=i, column=1))
                if val and 'Subject ID' in str(val):
                    subject_start_row = i + 1
                    print(f"[DEBUG] Found 'Subject ID' at row {i}, data starts at row {subject_start_row}")
                    break
            
            if not subject_start_row:
                return JsonResponse({'error': 'Could not find subject data.'}, status=400)
            
            # FIRST PASS: Collect all classes and check for conflicts
            classes_to_create = []
            conflicts = []
            row_idx = subject_start_row
            consecutive_empty = 0
            
            while True:
                subject_id = get_cell_value(sheet.cell(row=row_idx, column=1))
                subject_code = get_cell_value(sheet.cell(row=row_idx, column=2))
                description = get_cell_value(sheet.cell(row=row_idx, column=3))
                college = get_cell_value(sheet.cell(row=row_idx, column=4))
                time_from_val = get_cell_value(sheet.cell(row=row_idx, column=5))
                time_to_val = get_cell_value(sheet.cell(row=row_idx, column=6))
                day = get_cell_value(sheet.cell(row=row_idx, column=7))
                room = get_cell_value(sheet.cell(row=row_idx, column=8))
                program = get_cell_value(sheet.cell(row=row_idx, column=9))
                class_size_str = get_cell_value(sheet.cell(row=row_idx, column=10))
                section = get_cell_value(sheet.cell(row=row_idx, column=11))
                
                if subject_id and 'No.' in str(subject_id):
                    print(f"[DEBUG] Found student table at row {row_idx}")
                    break
                
                if not subject_id or str(subject_id).strip() == '':
                    consecutive_empty += 1
                    row_idx += 1
                    if consecutive_empty > 10:
                        break
                    continue
                else:
                    consecutive_empty = 0
                
                if not subject_code or str(subject_code).strip() == '':
                    row_idx += 1
                    continue
                
                time_from = parse_excel_time(time_from_val)
                time_to = parse_excel_time(time_to_val)
                
                if not time_from or not time_to:
                    row_idx += 1
                    continue
                
                # Check for schedule conflicts with existing classes
                day_str = str(day).strip() if day else ''
                time_from_str = time_from.strftime('%H:%M:%S')
                time_to_str = time_to.strftime('%H:%M:%S')
                
                # Find conflicting classes
                conflicting_classes = Class.objects.filter(
                    teacher=request.user,
                    day=day_str,
                    time_from__lt=time_to,
                    time_to__gt=time_from
                )
                
                for conflict in conflicting_classes:
                    conflicts.append({
                        'new_class': str(subject_code).strip(),
                        'new_schedule': f"{day_str} {time_from.strftime('%I:%M %p')} - {time_to.strftime('%I:%M %p')}",
                        'existing_class': f"{conflict.subject_code} - {conflict.subject_description}",
                        'existing_schedule': f"{conflict.day} {conflict.time_from.strftime('%I:%M %p')} - {conflict.time_to.strftime('%I:%M %p')}"
                    })
                
                classes_to_create.append({
                    'subject_code': str(subject_code).strip(),
                    'section': str(section).strip() if section else '',
                    'semester': str(semester).strip() if semester else '',
                    'school_year': str(school_year).strip() if school_year else '',
                    'day': day_str,
                    'time_from': time_from,
                    'time_to': time_to,
                    'subject_description': str(description).strip() if description else '',
                    'college': str(college).strip() if college else '',
                    'student_type': str(student_type).strip() if student_type else '',
                    'room': str(room).strip() if room else '',
                    'program': str(program).strip() if program else '',
                    'class_size': int(class_size_str) if class_size_str and str(class_size_str).isdigit() else 0,
                })
                
                row_idx += 1
                
                if row_idx > subject_start_row + 50:
                    break
            
            # If there are conflicts, return them to show in modal
            if conflicts:
                return JsonResponse({
                    'success': False,
                    'has_conflicts': True,
                    'conflicts': conflicts,
                    'message': f'Found {len(conflicts)} schedule conflict(s). Please review before uploading.'
                }, status=409)
            
            # SECOND PASS: No conflicts, proceed with creating classes
            classes_created = []
            for class_data in classes_to_create:
                class_obj, created = Class.objects.get_or_create(
                    subject_code=class_data['subject_code'],
                    section=class_data['section'],
                    semester=class_data['semester'],
                    school_year=class_data['school_year'],
                    day=class_data['day'],
                    time_from=class_data['time_from'],
                    time_to=class_data['time_to'],
                    defaults={
                        'subject_description': class_data['subject_description'],
                        'college': class_data['college'],
                        'student_type': class_data['student_type'],
                        'room': class_data['room'],
                        'program': class_data['program'],
                        'class_size': class_data['class_size'],
                        'teacher': request.user,
                    }
                )
                if created:
                    classes_created.append(class_obj)
            
            # Continue with student processing...
            student_start_row = None
            for i in range(1, 500):
                col1 = get_cell_value(sheet.cell(row=i, column=1))
                col2 = get_cell_value(sheet.cell(row=i, column=2))
                if col1 and 'No.' in str(col1) and col2 and 'Student ID' in str(col2):
                    student_start_row = i + 1
                    break
            
            if not student_start_row:
                if len(classes_created) > 0:
                    return JsonResponse({'success': True, 'message': f"Uploaded {len(classes_created)} class(es)."})
                else:
                    return JsonResponse({'error': 'Could not find student data'}, status=400)
            
            students_processed = 0
            students_skipped = []
            row_idx = student_start_row
            
            while True:
                student_id = get_cell_value(sheet.cell(row=row_idx, column=2))
                name = get_cell_value(sheet.cell(row=row_idx, column=3))
                email = get_cell_value(sheet.cell(row=row_idx, column=4))
                
                if not student_id or not name:
                    break
                
                if not email:
                    email = f"{student_id}@wmsu.edu.ph"
                
                name_parts = str(name).split(',')
                if len(name_parts) == 2:
                    last_name = name_parts[0].strip()
                    first_middle = name_parts[1].strip().split()
                    first_name = first_middle[0] if first_middle else ''
                else:
                    first_name = str(name)
                    last_name = ''
                
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'user_type': 'student',
                    }
                )
                if not created:
                    if hasattr(user, 'student_id') and user.student_id != student_id:
                        user.student_id = student_id
                        user.save()
                    students_skipped.append(f"{name} (already exists)")
                else:
                    if hasattr(user, 'student_id'):
                        user.student_id = student_id
                        user.save()
                    students_processed += 1
                
                for class_obj in classes_created:
                    Enrollment.objects.get_or_create(
                        student=user,
                        class_obj=class_obj,
                        defaults={'status': 'enrolled', 'absence_count': 0}
                    )
                
                row_idx += 1
                
                if row_idx > student_start_row + 500:
                    break
            
            if len(classes_created) == 0:
                return JsonResponse({'error': 'No valid classes found.'}, status=400)
            
            message = f"Uploaded {len(classes_created)} class(es) and enrolled {students_processed} student(s)."
            if students_skipped:
                message += f" Skipped {len(students_skipped)} existing students."
            
            return JsonResponse({'success': True, 'message': message})
        
        except Exception as e:
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'No file provided'}, status=400)


@csrf_exempt
@login_required
def delete_class(request, class_id):
    """Delete a class and its enrollments"""
    if request.method == 'DELETE':
        try:
            class_obj = Class.objects.get(id=class_id, teacher=request.user)
            class_name = str(class_obj)
            class_obj.delete()
            return JsonResponse({'success': True, 'message': f'Class "{class_name}" has been deleted.'})
        except Class.DoesNotExist:
            return JsonResponse({'error': 'Class not found.'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=400)


# ============================================
# CLASS SESSION AND ATTENDANCE VIEWS
# ============================================
@login_required
def start_class_session(request, class_id):
    """Start a new class session"""
    if request.method == 'POST':
        try:
            class_obj = Class.objects.get(id=class_id, teacher=request.user)
            print(f"[DEBUG] Starting class session for: {class_obj.subject_code}")
            
            active_session = ClassSession.objects.filter(class_obj=class_obj, status='active').first()
            if active_session:
                return JsonResponse({'error': 'A class session is already active for this class.'}, status=400)
            
            session = ClassSession.objects.create(
                class_obj=class_obj,
                teacher=request.user,
                status='active'
            )
            
            print(f"[DEBUG] Session created with ID: {session.id}")
            
            return JsonResponse({
                'success': True, 
                'message': 'Class session started successfully!',
                'session_id': session.id
            })
        except Class.DoesNotExist:
            return JsonResponse({'error': 'Class not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def end_class_session(request, session_id):
    """End an active class session and mark absent students"""
    if request.method == 'POST':
        try:
            session = ClassSession.objects.get(id=session_id, teacher=request.user, status='active')
            
            # Get all enrollments for this class
            enrollments = Enrollment.objects.filter(class_obj=session.class_obj).select_related('student')
            
            # Get all attendances for this session (current present students)
            attendances = Attendance.objects.filter(session=session).select_related('student')
            present_student_ids = [att.student.id for att in attendances]
            
            # Track students who were marked absent
            warning_students = []
            dropped_students = []
            newly_absent_students = []
            
            # Process each enrollment to mark absences
            for enrollment in enrollments:
                student = enrollment.student
                
                if student.id not in present_student_ids:
                    newly_absent_students.append(student)
                    old_absence_count = enrollment.absence_count
                    enrollment.absence_count += 1
                    absence_count = enrollment.absence_count
                    
                    if absence_count == 3 and old_absence_count < 3:
                        warning_students.append({
                            'student_name': student.get_full_name(),
                            'absence_count': absence_count,
                            'warning_level': 'first_warning',
                            'message': f'Warning: {absence_count}/5 absences. 2 remaining before being dropped.'
                        })
                    elif absence_count == 4 and old_absence_count < 4:
                        warning_students.append({
                            'student_name': student.get_full_name(),
                            'absence_count': absence_count,
                            'warning_level': 'final_warning',
                            'message': f'FINAL WARNING: {absence_count}/5 absences. 1 more absence and you will be dropped!'
                        })
                    elif absence_count >= 5 and enrollment.status != 'dropped':
                        enrollment.status = 'dropped'
                        dropped_students.append({
                            'student_name': student.get_full_name(),
                            'absence_count': absence_count,
                            'warning_level': 'dropped',
                            'message': f'DROPPED: Student has been dropped from the class due to {absence_count}/5 absences.'
                        })
                    
                    enrollment.save()
            
            # Create absent attendance records for absent students
            for student in newly_absent_students:
                Attendance.objects.get_or_create(
                    session=session,
                    student=student,
                    defaults={'status': 'absent'}
                )
            
            # CRITICAL: Set session status to 'ended'
            session.status = 'ended'
            session.end_time = timezone.now()
            session.save()
            
            print(f"[DEBUG] Session {session.id} ended with status: {session.status}")
            print(f"[DEBUG] Total attendances: {Attendance.objects.filter(session=session).count()}")
            
            # Prepare response
            all_attendances = Attendance.objects.filter(session=session)
            present_count = all_attendances.filter(status='present').count()
            absent_count = all_attendances.filter(status='absent').count()
            
            # Prepare attendance data for PDF (always save to database)
            attendance_data = []
            for enrollment in enrollments:
                student = enrollment.student
                attendance = all_attendances.filter(student=student).first()
                
                is_present = attendance and attendance.status == 'present'
                
                attendance_data.append({
                    'student_id': student.student_id if hasattr(student, 'student_id') else 'N/A',
                    'name': student.get_full_name(),
                    'email': student.email,
                    'course': getattr(student, 'course', 'N/A'),
                    'status': 'Present' if is_present else 'Absent',
                    'time_in': attendance.time_in.strftime('%I:%M %p') if attendance and attendance.time_in and is_present else '—'
                })
            
            attendance_data.sort(key=lambda x: (x['status'] != 'Present', x['name']))
            
            # ALWAYS generate and save PDF to database (regardless of download option)
            try:
                from django.core.files.base import ContentFile
                
                # Generate PDF content
                pdf_content = generate_attendance_pdf(
                    session.class_obj, 
                    session, 
                    attendance_data, 
                    present_count, 
                    absent_count
                )
                
                # Create filename
                filename = f"attendance_{session.class_obj.subject_code}_{session.start_time.strftime('%Y%m%d_%H%M%S')}.pdf"
                
                # Check if PDF already exists for this session
                existing_report = ClassPDFReport.objects.filter(session=session).first()
                
                if existing_report:
                    # Update existing report
                    existing_report.filename = filename
                    existing_report.file_size = len(pdf_content)
                    existing_report.total_students = len(attendance_data)
                    existing_report.present_count = present_count
                    existing_report.absent_count = absent_count
                    existing_report.attendance_rate = int((present_count/len(attendance_data))*100) if len(attendance_data) > 0 else 0
                    
                    # Delete old file and save new one
                    if existing_report.pdf_file:
                        existing_report.pdf_file.delete()
                    existing_report.pdf_file.save(filename, ContentFile(pdf_content))
                    existing_report.save()
                    print(f"✅ PDF updated in database: {filename}")
                else:
                    # Create new PDF report record
                    pdf_report = ClassPDFReport.objects.create(
                        session=session,
                        class_obj=session.class_obj,
                        teacher=request.user,
                        filename=filename,
                        file_size=len(pdf_content),
                        total_students=len(attendance_data),
                        present_count=present_count,
                        absent_count=absent_count,
                        attendance_rate=int((present_count/len(attendance_data))*100) if len(attendance_data) > 0 else 0
                    )
                    
                    # Save the PDF file
                    pdf_report.pdf_file.save(filename, ContentFile(pdf_content))
                    print(f"✅ PDF saved to database: {filename}")
                
            except Exception as e:
                print(f"Error saving PDF to database: {e}")
                import traceback
                traceback.print_exc()
            
            # Check if PDF download is requested
            generate_pdf = request.POST.get('generate_pdf', False) or request.GET.get('pdf', False)
            
            if generate_pdf:
                # Return PDF for download
                response = HttpResponse(pdf_content, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="Attendance_{session.class_obj.subject_code}_{session.start_time.strftime("%Y%m%d_%H%M")}.pdf"'
                return response
            
            # Return JSON for AJAX request (no PDF download)
            return JsonResponse({
                'success': True,
                'present_count': present_count,
                'absent_count': absent_count,
                'total_students': enrollments.count(),
                'warnings': warning_students,
                'dropped': dropped_students,
                'pdf_saved': True,
                'message': 'Class session ended and attendance report saved to history.'
            })
            
        except ClassSession.DoesNotExist:
            return JsonResponse({'error': 'Active session not found'}, status=404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def active_class_session(request, class_id):
    """Get active session for a class and display attendance page"""
    try:
        class_obj = Class.objects.get(id=class_id, teacher=request.user)
        active_session = ClassSession.objects.filter(class_obj=class_obj, status='active').first()
        
        if not active_session:
            messages.error(request, 'No active session for this class.')
            return redirect(f'/view_class/?class_id={class_id}')
        
        attendances = Attendance.objects.filter(session=active_session).select_related('student')
        enrollments = Enrollment.objects.filter(class_obj=class_obj).select_related('student')
        
        students_data = []
        for enrollment in enrollments:
            attendance = attendances.filter(student=enrollment.student).first()
            students_data.append({
                'id': enrollment.student.id,
                'name': enrollment.student.get_full_name(),
                'student_id': enrollment.student.student_id if hasattr(enrollment.student, 'student_id') else '',
                'email': enrollment.student.email,
                'status': attendance.status if attendance else 'Not yet tapped',
                'time_in': attendance.time_in.strftime('%I:%M %p') if attendance and attendance.time_in else None,
            })
        
        attendance_rate = int((len(attendances) / len(students_data)) * 100) if len(students_data) > 0 else 0
        
        context = {
            'class': class_obj,
            'session': active_session,
            'students': students_data,
            'total_students': len(students_data),
            'present_count': len(attendances),
            'attendance_rate': attendance_rate,
        }
        return render(request, 'active_class_session.html', context)
        
    except Class.DoesNotExist:
        messages.error(request, 'Class not found.')
        return redirect('classes')


@csrf_exempt
def record_attendance(request):
    """Record student attendance via RFID scanner - NO LOGIN REQUIRED"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rfid_tag = data.get('rfid_tag')
            session_id = data.get('session_id')
            
            if not rfid_tag or not session_id:
                return JsonResponse({'error': 'Missing RFID tag or session ID'}, status=400)
            
            session = ClassSession.objects.get(id=session_id, status='active')
            
            student = User.objects.filter(rfid_tag=rfid_tag, user_type='student').first()
            if not student:
                return JsonResponse({'error': 'Student not found. Please register your RFID first.'}, status=404)
            
            enrollment = Enrollment.objects.filter(student=student, class_obj=session.class_obj).first()
            if not enrollment:
                return JsonResponse({'error': f'{student.get_full_name()} is not enrolled in this class.'}, status=403)
            
            if enrollment.status == 'dropped':
                return JsonResponse({'error': f'{student.get_full_name()} has been dropped from this class.'}, status=403)
            
            existing_attendance = Attendance.objects.filter(session=session, student=student).first()
            if existing_attendance:
                return JsonResponse({'error': f'{student.get_full_name()} already recorded attendance.'}, status=400)
            
            # Create attendance - time will be set automatically with local time
            attendance = Attendance.objects.create(
                session=session,
                student=student,
                status='present'
            )
            
            # Format time for response
            time_formatted = attendance.time_in.strftime('%I:%M %p')
            
            return JsonResponse({
                'success': True,
                'message': f'✅ Attendance recorded for {student.get_full_name()}',
                'student_name': student.get_full_name(),
                'time_in': time_formatted
            })
            
        except ClassSession.DoesNotExist:
            return JsonResponse({'error': 'No active session found. Teacher needs to start the class first.'}, status=404)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def get_active_session(request):
    """API endpoint for RFID reader to check active session - NO LOGIN REQUIRED"""
    active_session = ClassSession.objects.filter(status='active').first()
    
    if active_session:
        return JsonResponse({
            'has_active_session': True,
            'session_id': active_session.id,
            'class_name': active_session.class_obj.subject_description,
        })
    else:
        return JsonResponse({'has_active_session': False})


@login_required
def activity(request):
    return render(request, 'activity.html')


@login_required
def schedule(request):
    return render(request, 'schedule.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('adminLogin')


def generate_attendance_pdf(class_obj, session, attendance_data, present_count, absent_count):
    """Generate PDF report for attendance"""
    
    # Create buffer for PDF
    buffer = io.BytesIO()
    
    # Create PDF document (landscape for better table display)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), 
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=72)
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#5A1219'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        textColor=colors.grey,
        spaceAfter=20
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.white,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph(f"Attendance Report - {class_obj.subject_description}", title_style)
    elements.append(title)
    
    # Class details
    class_details = f"""
    <b>Subject Code:</b> {class_obj.subject_code}<br/>
    <b>Section:</b> {class_obj.section if class_obj.section else 'N/A'}<br/>
    <b>Room:</b> {class_obj.room if class_obj.room else 'TBA'}<br/>
    <b>Schedule:</b> {class_obj.day if class_obj.day else 'TBA'} | {class_obj.time_from.strftime('%I:%M %p') if class_obj.time_from else 'N/A'} - {class_obj.time_to.strftime('%I:%M %p') if class_obj.time_to else 'N/A'}<br/>
    <b>School Year:</b> {class_obj.school_year if class_obj.school_year else 'N/A'}<br/>
    <b>Semester:</b> {class_obj.semester if class_obj.semester else 'N/A'}
    """
    
    details_paragraph = Paragraph(class_details, subtitle_style)
    elements.append(details_paragraph)
    elements.append(Spacer(1, 10))
    
    # Session info
    session_info = f"""
    <b>Session Date:</b> {session.start_time.strftime('%B %d, %Y')}<br/>
    <b>Start Time:</b> {session.start_time.strftime('%I:%M %p')}<br/>
    <b>End Time:</b> {session.end_time.strftime('%I:%M %p') if session.end_time else 'Ongoing'}<br/>
    <b>Total Students:</b> {len(attendance_data)}<br/>
    <b>Present:</b> {present_count} ({int(present_count/len(attendance_data)*100) if len(attendance_data) > 0 else 0}%)<br/>
    <b>Absent:</b> {absent_count} ({int(absent_count/len(attendance_data)*100) if len(attendance_data) > 0 else 0}%)
    """
    
    session_paragraph = Paragraph(session_info, subtitle_style)
    elements.append(session_paragraph)
    elements.append(Spacer(1, 20))
    
    # Prepare table data
    table_data = []
    
    # Table header
    headers = ['#', 'Student ID', 'Student Name', 'Course', 'Status', 'Time In']
    table_data.append([Paragraph(h, header_style) for h in headers])
    
    # Table rows
    for idx, student in enumerate(attendance_data, 1):
        row = [
            str(idx),
            student.get('student_id', 'N/A'),
            student.get('name', 'N/A'),
            student.get('course', 'N/A'),
            student.get('status', 'Absent'),
            student.get('time_in', '—')
        ]
        table_data.append(row)
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Table styling
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5A1219')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Body styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        
        # Grid styling
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#D1D5DB')),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))
    
    # Color code status column based on value
    for i, row in enumerate(table_data[1:], start=1):
        status = row[4]
        if 'present' in status.lower():
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#059669')),
                ('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'),
            ]))
        elif 'absent' in status.lower():
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (4, i), (4, i), colors.HexColor('#DC2626')),
                ('FONTNAME', (4, i), (4, i), 'Helvetica-Bold'),
            ]))
    
    elements.append(table)
    
    # Add footer with generation timestamp
    elements.append(Spacer(1, 30))
    footer_text = f"Report generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Get PDF content
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


# ============================================
# ATTENDANCE HISTORY API
# ============================================
@login_required
@require_http_methods(["GET"])
def get_student_attendance_history(request, student_id, class_id):
    """API endpoint to get attendance history for a specific student in a specific class"""
    try:
        # Verify teacher owns this class
        class_obj = Class.objects.get(id=class_id, teacher=request.user)
        
        # Get the student
        student = User.objects.get(id=student_id, user_type='student')
        
        # Get all ended sessions for this class, ordered by most recent first
        sessions = ClassSession.objects.filter(
            class_obj=class_obj,
            status='ended'
        ).order_by('-start_time')[:10]
        
        attendance_history = []
        
        for session in sessions:
            # Check if student has attendance record for this session
            attendance = Attendance.objects.filter(session=session, student=student).first()
            
            if attendance:
                attendance_history.append({
                    'date': session.start_time.strftime('%B %d, %Y'),
                    'time': session.start_time.strftime('%I:%M %p'),
                    'status': attendance.status,
                    'time_in': attendance.time_in.strftime('%I:%M %p') if attendance.time_in else None
                })
            else:
                # Student was absent (no attendance record for ended session)
                attendance_history.append({
                    'date': session.start_time.strftime('%B %d, %Y'),
                    'time': session.start_time.strftime('%I:%M %p'),
                    'status': 'absent',
                    'time_in': None
                })
        
        return JsonResponse({
            'success': True,
            'attendance_history': attendance_history,
            'total_sessions': len(attendance_history)
        })
        
    except Class.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Class not found'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student not found'}, status=404)
    except Exception as e:
        print(f"Error in get_student_attendance_history: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def get_activity_log(request):
    """API endpoint to get activity log for the teacher's classes"""
    try:
        # Get all attendance records for classes taught by this teacher
        # Get all classes taught by the teacher
        teacher_classes = Class.objects.filter(teacher=request.user)
        
        # Get all sessions for these classes
        sessions = ClassSession.objects.filter(
            class_obj__in=teacher_classes
        ).order_by('-start_time')
        
        # Build activity log data
        activities = []
        for session in sessions:
            attendances = Attendance.objects.filter(session=session).select_related('student')
            
            for attendance in attendances:
                activities.append({
                    'id': attendance.id,
                    'student_name': attendance.student.get_full_name(),
                    'student_id': attendance.student.student_id or 'N/A',
                    'class_name': session.class_obj.subject_code,
                    'class_description': session.class_obj.subject_description,
                    'section': session.class_obj.section or 'N/A',
                    'date': session.start_time.strftime('%B %d, %Y'),
                    'time': session.start_time.strftime('%I:%M %p'),
                    'status': attendance.status,
                    'time_in': attendance.time_in.strftime('%I:%M %p') if attendance.time_in else 'N/A'
                })
        
        return JsonResponse({
            'success': True,
            'activities': activities
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def update_attendance_status(request):
    """API endpoint to update attendance status"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            attendance_id = data.get('attendance_id')
            new_status = data.get('status')
            
            attendance = Attendance.objects.get(id=attendance_id)
            
            # Verify teacher owns this class
            if attendance.session.class_obj.teacher != request.user:
                return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)
            
            # Update status
            attendance.status = new_status
            attendance.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Status updated to {new_status}'
            })
            
        except Attendance.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Attendance record not found'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=400)


@login_required
def get_session_attendance_api(request, session_id):
    """API endpoint to get attendance updates for a session"""
    try:
        session = ClassSession.objects.get(id=session_id)
        attendances = Attendance.objects.filter(session=session).select_related('student')
        
        attendance_list = []
        for attendance in attendances:
            # Convert to local time for display
            local_time = attendance.time_in
            attendance_list.append({
                'student_id': attendance.student.id,
                'student_name': attendance.student.get_full_name(),
                'time_in': local_time.strftime('%I:%M %p'),
            })
        
        return JsonResponse({
            'success': True,
            'present_count': attendances.count(),
            'attendance_rate': int((attendances.count() / session.class_obj.total_students) * 100) if session.class_obj.total_students > 0 else 0,
            'attendance_list': attendance_list
        })
    except ClassSession.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Session not found'}, status=404)


# ============================================
# PDF REPORT MANAGEMENT VIEWS
# ============================================
@login_required
def view_pdf_reports(request, class_id):
    """View all PDF reports for a specific class"""
    try:
        class_obj = Class.objects.get(id=class_id, teacher=request.user)
        reports = ClassPDFReport.objects.filter(class_obj=class_obj, teacher=request.user)
        
        context = {
            'class': class_obj,
            'reports': reports,
            'total_reports': reports.count()
        }
        return render(request, 'pdf_reports.html', context)
    except Class.DoesNotExist:
        messages.error(request, 'Class not found')
        return redirect('classes')


@login_required
def download_pdf_report(request, report_id):
    """Download a specific PDF report"""
    try:
        report = ClassPDFReport.objects.get(id=report_id, teacher=request.user)
        
        if not report.pdf_file:
            return JsonResponse({'error': 'PDF file not found'}, status=404)
        
        # Serve the file
        response = FileResponse(report.pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{report.filename}"'
        return response
        
    except ClassPDFReport.DoesNotExist:
        raise Http404("PDF report not found")


@login_required
def delete_pdf_report(request, report_id):
    """Delete a PDF report"""
    if request.method == 'DELETE':
        try:
            report = ClassPDFReport.objects.get(id=report_id, teacher=request.user)
            # Delete the physical file
            if report.pdf_file:
                report.pdf_file.delete()
            report.delete()
            return JsonResponse({'success': True, 'message': 'PDF report deleted successfully'})
        except ClassPDFReport.DoesNotExist:
            return JsonResponse({'error': 'PDF report not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=400)